"""Generate a blank xlsx template with headers + 1 example row + cell comments.

Output: examples/blank-template.xlsx

Usage:
    python examples/generate_blank_template.py
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "blank-template.xlsx"

HEADERS = [
    ("ID", 12, "Numeric ID for the test case. Required for Test Case rows. Leave blank on step rows."),
    ("Work Item Type", 18, "'Test Case' opens a new case. Other rows (or empty) are steps belonging to the previous case. 'Shared Steps' is treated like a step row."),
    ("Title", 50, "Test case title. Required for Test Case rows. Leave blank on step rows."),
    ("Test Step", 12, "Step number (e.g. 1, 2, 2.1). Optional but recommended."),
    ("Step Action", 60, "What the tester does in this step. Imperative voice ('Click X', 'Enter Y'). Required for step rows."),
    ("Step Expected", 60, "Expected outcome after the action. Optional but encouraged."),
    ("Area Path", 30, "Project / area classification. Optional. Free text."),
    ("Assigned To", 30, "Tester or owner. Optional. Free text or email."),
    ("State", 14, "Workflow state, e.g. 'Design', 'Ready', 'In Progress', 'Closed'. Optional."),
]

EXAMPLE = [
    [101, "Test Case", "Login - successful sign-in with valid credentials",
     None, None, None, "MyApp\\Auth", "alice@example.com", "Design"],
    [None, None, None, "1", "Open the application URL", "Login page is displayed", None, None, None],
    [None, None, None, "2", "Type a valid username in the Email field", "Field accepts the value", None, None, None],
    [None, None, None, "3", "Type a valid password in the Password field", "Characters are masked", None, None, None],
    [None, None, None, "4", "Click the Sign In button", "User is redirected to the dashboard", None, None, None],
    [102, "Test Case", "Login - failure with invalid password",
     None, None, None, "MyApp\\Auth", "alice@example.com", "Design"],
    [None, None, None, "1", "Open the application URL", "Login page is displayed", None, None, None],
    [None, None, None, "2", "Type a valid username", "Field accepts the value", None, None, None],
    [None, None, None, "3", "Type an invalid password", "Characters are masked", None, None, None],
    [None, None, None, "4", "Click the Sign In button",
     "Error message 'Invalid credentials' is displayed", None, None, None],
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    border = Border(left=Side(style="thin", color="999999"),
                    right=Side(style="thin", color="999999"),
                    top=Side(style="thin", color="999999"),
                    bottom=Side(style="thin", color="999999"))
    case_fill = PatternFill("solid", fgColor="DDEBF7")

    # Header row with comments
    for i, (name, width, comment_text) in enumerate(HEADERS, 1):
        col = get_column_letter(i)
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = border
        cell.comment = Comment(comment_text, "tfs-test-runner")
        cell.comment.width = 320
        cell.comment.height = 110
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Example rows
    for r, row in enumerate(EXAMPLE, 2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row[1] == "Test Case":
                cell.fill = case_fill
                cell.font = Font(bold=True)
        ws.row_dimensions[r].height = 30 if row[1] == "Test Case" else 24

    # Add a few empty rows hinting where to start filling
    hint_row = len(EXAMPLE) + 2
    hint = ws.cell(row=hint_row, column=1, value="↓ Add your own test cases below ↓")
    hint.font = Font(italic=True, color="888888")
    hint.alignment = Alignment(horizontal="left")

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()

"""Parse xlsx test cases from Azure DevOps export.

Detects columns by header (case-insensitive). Returns list of cases:
[{id, title, assigned, state, area, steps:[{step, action, expected, shared_id}]}, ...]
"""
from __future__ import annotations
import re, unicodedata
from pathlib import Path
from typing import Any
from openpyxl import load_workbook


HEADER_ALIASES = {
    "id": ["id", "work item id"],
    "type": ["work item type", "type"],
    "title": ["title", "name"],
    "step": ["test step", "step"],
    "action": ["step action", "action"],
    "expected": ["step expected", "expected", "step expected result"],
    "area": ["area path", "area"],
    "assigned": ["assigned to", "assignee"],
    "state": ["state", "status"],
}


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def _find_columns(header_row: list) -> dict[str, int]:
    norm = [_norm_header(c) for c in header_row]
    out: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        for i, h in enumerate(norm):
            if h in aliases:
                out[key] = i
                break
    missing = [k for k in ("id", "type", "title", "action") if k not in out]
    if missing:
        raise ValueError(f"Spreadsheet missing required columns: {missing}. Found: {norm}")
    return out


def _clean(text: Any) -> str:
    if text is None:
        return ""
    s = str(text)
    s = s.replace("\xa0", " ").replace("​", "")
    s = unicodedata.normalize("NFC", s)
    return s.strip()


def parse_xlsx(path: str | Path) -> list[dict]:
    """Read xlsx and return list of test cases with steps."""
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    cols = _find_columns(list(header))

    def _opt(row, key):
        idx = cols.get(key)
        return _clean(row[idx]) if idx is not None else ""

    cases: list[dict] = []
    current: dict | None = None

    for row in rows:
        # iter_rows always returns tuples; skip rows that are entirely None
        if not any(c is not None for c in row):
            continue
        # required columns guaranteed present by _find_columns
        wtype = _clean(row[cols["type"]])
        wid = _clean(row[cols["id"]])
        title = _clean(row[cols["title"]])
        action = _clean(row[cols["action"]])
        # optional
        step = _opt(row, "step")
        expected = _opt(row, "expected")
        area = _opt(row, "area")
        assigned = _opt(row, "assigned")
        state = _opt(row, "state")

        if wtype.lower() == "test case":
            current = {
                "id": wid,
                "title": title,
                "assigned": assigned,
                "state": state,
                "area": area,
                "steps": [],
            }
            cases.append(current)
        elif current is not None:
            if step or action or expected:
                current["steps"].append({
                    "step": step,
                    "action": action,
                    "expected": expected,
                    "shared_id": wid if wtype.lower() == "shared steps" else "",
                })

    wb.close()
    return cases


if __name__ == "__main__":
    import sys, json
    cases = parse_xlsx(sys.argv[1])
    print(f"cases: {len(cases)}", file=sys.stderr)
    print(f"steps: {sum(len(c['steps']) for c in cases)}", file=sys.stderr)
    print(json.dumps(cases, ensure_ascii=False, indent=2))

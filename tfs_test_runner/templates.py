"""Bundled templates for `tfs-test-runner init`.

Generates a blank xlsx, phases YAML, and glossary YAML at runtime so the
package has no .xlsx asset to ship (keeps the wheel small and avoids
gitignore juggling).
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


_HEADERS = [
    ("ID", 12, "Numeric ID for the test case. Required for Test Case rows. Leave blank on step rows."),
    ("Work Item Type", 18, "'Test Case' opens a new case. Other rows (or empty) are steps belonging to the previous case."),
    ("Title", 50, "Test case title. Required for Test Case rows. Leave blank on step rows."),
    ("Test Step", 12, "Step number (e.g. 1, 2, 2.1). Optional but recommended."),
    ("Step Action", 60, "What the tester does. Imperative voice. Required for step rows."),
    ("Step Expected", 60, "Expected outcome after the action. Optional but encouraged."),
    ("Area Path", 30, "Project / area classification. Optional. Free text."),
    ("Assigned To", 30, "Tester or owner. Optional."),
    ("State", 14, "Workflow state, e.g. 'Design', 'Ready'. Optional."),
]

_EXAMPLE = [
    [101, "Test Case", "Login - successful sign-in with valid credentials",
     None, None, None, "MyApp\\Auth", "alice@example.com", "Design"],
    [None, None, None, "1", "Open the application URL", "Login page is displayed", None, None, None],
    [None, None, None, "2", "Type a valid username in the Email field", "Field accepts the value", None, None, None],
    [None, None, None, "3", "Type a valid password in the Password field", "Characters are masked", None, None, None],
    [None, None, None, "4", "Click the Sign In button", "User is redirected to the dashboard", None, None, None],
    [102, "Test Case", "Login - failure with invalid password",
     None, None, None, "MyApp\\Auth", "alice@example.com", "Design"],
    [None, None, None, "1", "Open the application URL", "Login page is displayed", None, None, None],
    [None, None, None, "2", "Type a valid username", "Field accepts the value", None, None, None],
    [None, None, None, "3", "Type an invalid password", "Characters are masked", None, None, None],
    [None, None, None, "4", "Click the Sign In button",
     "Error message 'Invalid credentials' is displayed", None, None, None],
]


def write_blank_xlsx(path: str | Path) -> Path:
    """Create a blank test-case xlsx with header row + 2 example test cases."""
    out = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    border = Border(left=Side(style="thin", color="999999"),
                    right=Side(style="thin", color="999999"),
                    top=Side(style="thin", color="999999"),
                    bottom=Side(style="thin", color="999999"))
    case_fill = PatternFill("solid", fgColor="DDEBF7")

    for i, (name, width, comment_text) in enumerate(_HEADERS, 1):
        col = get_column_letter(i)
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = border
        cell.comment = Comment(comment_text, "tfs-test-runner")
        cell.comment.width = 320
        cell.comment.height = 110
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for r, row in enumerate(_EXAMPLE, 2):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row[1] == "Test Case":
                cell.fill = case_fill
                cell.font = Font(bold=True)
        ws.row_dimensions[r].height = 30 if row[1] == "Test Case" else 24

    hint_row = len(_EXAMPLE) + 2
    hint = ws.cell(row=hint_row, column=1, value="Add your own test cases below")
    hint.font = Font(italic=True, color="888888")

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


PHASES_YAML = """\
# Group test cases into phases for organized execution.
# Without this file, all cases land in a single "All cases" phase.

phases:
  - id: "p1"
    title: "Phase 1 — Smoke / happy path"
    level: easy            # easy | med | hard (drives badge color)
    desc: "Critical positive scenarios run on every build."
    case_ids: ["101"]

  - id: "p2"
    title: "Phase 2 — Failure scenarios"
    level: med
    desc: "Negative paths and error handling."
    match: ["failure", "invalid", "error"]    # case-insensitive substring on title
"""


GLOSSARY_YAML = """\
# Glossary refines LLM translation behavior. Used only with --llm.
#
# preserve: list of terms the model must keep verbatim (UI labels, tech terms).
# notes:    free-text appended to the system prompt for domain context.

preserve:
  - "Sign In"
  - "Save"
  - "Cancel"
  - "Profile"
  - "Email"
  - "Password"

notes: |
  Domain: web application QA testing. Users execute test cases manually.
  Tone: technical, imperative ("Click", "Type", "Verify").
"""


INIT_README = """# QA test plan

Generated by `tfs-test-runner init`.

## Files

- `cases.xlsx` — fill with your test cases. Schema: `ID`, `Work Item Type`,
  `Title`, `Test Step`, `Step Action`, `Step Expected`, `Area Path`,
  `Assigned To`, `State`. Hover the header cells for column comments.
- `phases.yaml` — group cases into phases (optional).
- `glossary.yaml` — refine LLM translation (optional, only used with `--llm`).

## Generate the HTML kit

```bash
tfs-test-runner plan cases.xlsx -o plan.html
# with phases:
tfs-test-runner plan cases.xlsx --phases phases.yaml -o plan.html
# with GPT translation (needs OPENAI_API_KEY):
tfs-test-runner plan cases.xlsx --llm --lang pt-BR -o plan.html
```

Open `plan.html` in any browser to start executing tests.

## Validate the xlsx

```bash
tfs-test-runner validate cases.xlsx
```

Documentation: https://luizhcrs.github.io/tfs-test-runner/
"""
